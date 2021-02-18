from openpyxl import Workbook
from openpyxl import load_workbook
import re

wb = load_workbook('/Users/jichen/Desktop/workshift spreadsheet.xlsx')
print(wb.get_sheet_names())

ws = wb['zhang yu']
look_up_w = 5
saver = []
counter = 0
for row in ws.iter_rows(min_col = look_up_w, max_col = 7 + look_up_w, values_only = True):
    saver.append(row)
    counter += 1
print(counter)
print(len(saver))
print(saver[0])
saver.pop(0)
print(len(saver))
print(saver[0])

Empty_slot = 12

for i in range(0, len(saver), 2):
    print('Person:'+ str(i))
    #'-7'是判断点，第7周进项目
    if re.search(r'-7',str(saver[i][0])) != None:
        #’-7‘后有任何非空需要记作占用
        for k in range(1,len(saver[i])):
            if saver[i][k] != None:
                Empty_slot -= 1
                break
        #语句不执行则全是空
        print("Line 1:"+str(k))
        if k == len(saver[i])-1:
            for j in range(len(saver[i+1])):
                if saver[i+1][j] != None:
                    Empty_slot -= 1
                    break
    #如果第二行有'-7'
    elif re.search(r'-7',str(saver[i+1][0])) != None:
        for k in range(1,len(saver[i+1])):
            if saver[i+1][k] != None:
                Empty_slot -= 1
                break
        #语句不执行则全是空
        print("Line 2:"+str(k))
        if k == len(saver[i+1])-1:
            for j in range(len(saver[i])):
                if saver[i][j] != None:
                    Empty_slot -= 1
                    break 
    else:
        for k in range(len(saver[i])):
            if saver[i][k] != None:
                Empty_slot -= 1
                break
        if k == len(saver[i])-1:
            for j in range(len(saver[i+1])):
                if saver[i+1][j] != None:
                    Empty_slot -= 1
                    break
    print('-----Checked-----')
print(Empty_slot)